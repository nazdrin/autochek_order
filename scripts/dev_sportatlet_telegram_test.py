from __future__ import annotations

import os
import sys
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

import orchestrator  # noqa: E402
from supplier43_sportatlet_telegram import (  # noqa: E402
    fill_sportatlet_quantities,
    parse_sportatlet_items,
)


def _sample_order() -> dict:
    return {
        "id": 43001,
        "supplierlist": 43,
        "organizationId": 2,
        "ord_delivery_data": [{"trackingNumber": "20451398822945"}],
        "products": [
            {
                "id": "p1",
                "description": "CN3876, sport item",
                "amount": "2",
                "name": "Sport item 1",
                "costPerItem": "100",
            },
            {
                "id": "p2",
                "description": "AB-200 another item",
                "amount": "3",
                "name": "Sport item 2",
                "costPerItem": "200",
            },
            {
                "id": "p3",
                "description": "CN3876 duplicate",
                "amount": "1",
                "name": "Sport item 1 duplicate",
                "costPerItem": "100",
            },
        ],
    }


def _template_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.cell(6, 1).value = "Артикул"
    ws.cell(6, 2).value = "Кіл-ст"
    ws.cell(6, 3).value = "Ціна з макс. Знижкою"
    ws.cell(7, 1).value = "CN3876"
    ws.cell(7, 2).value = 0
    ws.cell(7, 3).value = 10.5
    ws.cell(8, 1).value = "AB-200"
    ws.cell(8, 2).value = 0
    ws.cell(8, 3).value = 20
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def main() -> int:
    order = _sample_order()
    items = parse_sportatlet_items(order)
    assert items == [
        {"article": "CN3876", "qty": 3, "name": "Sport item 1"},
        {"article": "AB-200", "qty": 3, "name": "Sport item 2"},
    ], items
    print("Sport-atlet item parsing ok")

    out_bytes, price_map = fill_sportatlet_quantities(_template_bytes(), items)
    wb = load_workbook(BytesIO(out_bytes), data_only=False)
    ws = wb.active
    assert ws.cell(7, 2).value == 3, ws.cell(7, 2).value
    assert ws.cell(8, 2).value == 3, ws.cell(8, 2).value
    assert price_map == {"CN3876": "10.50", "AB-200": "20.00"}, price_map
    print("Sport-atlet XLSX quantity fill ok")

    try:
        fill_sportatlet_quantities(_template_bytes(), [{"article": "MISSING", "qty": 1, "name": ""}])
    except RuntimeError as e:
        assert "MISSING" in str(e), str(e)
        print("Sport-atlet missing article failure ok")
    else:
        raise AssertionError("Expected missing article failure")

    old_env = {
        "BIOTUS_NP_API_KEY": os.environ.get("BIOTUS_NP_API_KEY"),
        "NP_API_KEY": os.environ.get("NP_API_KEY"),
        "BIOTUS_NP_API_KEY_ORG_2": os.environ.get("BIOTUS_NP_API_KEY_ORG_2"),
        "NP_API_KEY_ORG_2": os.environ.get("NP_API_KEY_ORG_2"),
    }
    old_salesdrive_update_status = orchestrator.salesdrive_update_status
    try:
        os.environ["BIOTUS_NP_API_KEY"] = "main-key"
        os.environ.pop("NP_API_KEY", None)
        os.environ["BIOTUS_NP_API_KEY_ORG_2"] = "org2-key"
        os.environ.pop("NP_API_KEY_ORG_2", None)
        key, source = orchestrator.resolve_np_api_key_for_order(order)
        assert key == "org2-key", (key, source)
        assert source == "BIOTUS_NP_API_KEY_ORG_2", source
        print("Sport-atlet NP organization key selection ok")

        updates: list[tuple[int, int, str, int]] = []

        def fake_salesdrive_update_status(order_id, status_id, comment=None, number_sup=None, products=None):
            updates.append((int(order_id), int(status_id), str(number_sup or ""), len(products or [])))

        orchestrator.salesdrive_update_status = fake_salesdrive_update_status
        state = {
            "sportatlet_sent": {
                "43001": {
                    "ttn": "20451398822945",
                    "chat_id": "-100123",
                }
            }
        }
        processed = orchestrator.process_one_order(order, state=state)
        assert processed is True
        assert updates == [(43001, orchestrator.ORCH_DONE_STATUS_ID, "TG_SENT", 3)], updates
        print("Sport-atlet idempotent status retry ok")
    finally:
        orchestrator.salesdrive_update_status = old_salesdrive_update_status
        for key, value in old_env.items():
            if value is None:
                os.environ.pop(key, None)
            else:
                os.environ[key] = value

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
