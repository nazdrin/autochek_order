from __future__ import annotations

import json
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple

from dotenv import load_dotenv
from playwright.sync_api import TimeoutError as PWTimeoutError
from playwright.sync_api import sync_playwright

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

load_dotenv(ROOT / ".env")

from services.gdrive_uploader import upload_or_update_json


def _env(name: str, default: str = "") -> str:
    return (os.getenv(name) or default).strip()


def _to_bool(raw: str, default: bool = False) -> bool:
    s = (raw or "").strip().lower()
    if not s:
        return default
    return s in {"1", "true", "yes", "y", "on"}


def _to_int(raw: str, default: int) -> int:
    try:
        val = int((raw or "").strip())
        if val > 0:
            return val
    except Exception:
        pass
    return default


SUP2_EXPORT_URL = _env("SUP2_EXPORT_URL", "https://crm.dobavki.ua/client/product/list/")
SUP2_EXPORT_DIR = _env("SUP2_EXPORT_DIR", "exports/dobavki")
SUP2_EXPORT_WAIT_SEC = _to_int(_env("SUP2_EXPORT_WAIT_SEC", "150"), 150)
SUP2_EXPORT_JSON_NAME = _env("SUP2_EXPORT_JSON_NAME", "dobavki_products.json")

SUP2_LOGIN_URL = _env("SUP2_LOGIN_URL", "https://crm.dobavki.ua/client/login")
SUP2_USERNAME = _env("SUP2_USERNAME")
SUP2_PASSWORD = _env("SUP2_PASSWORD")
SUP2_STORAGE_STATE_FILE = _env("SUP2_STORAGE_STATE_FILE", ".state_supplier2.json")
SUP2_TIMEOUT_MS = _to_int(_env("SUP2_TIMEOUT_MS", "20000"), 20000)
SUP2_HEADLESS = _to_bool(_env("SUP2_HEADLESS", "1"), True)

GDRIVE_FOLDER_ID = _env("GDRIVE_FOLDER_ID")
GDRIVE_CREDENTIALS_FILE = _env("GDRIVE_CREDENTIALS_FILE", "credentials.json")


def _norm_header(value: Any) -> str:
    s = str(value or "").strip().lower()
    s = s.replace("`", "")
    s = re.sub(r"[\s\-_./:;()\[\]{}]+", "", s)
    return s


def _extract_int_qty(value: Any) -> int:
    if value is None:
        return 0
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        try:
            return max(0, int(value))
        except Exception:
            return 0
    text = str(value).strip()
    if not text:
        return 0
    m_plus = re.search(r"(\d+)\s*\+", text)
    if m_plus:
        return int(m_plus.group(1))
    m_num = re.search(r"(\d+)", text)
    if m_num:
        return int(m_num.group(1))
    return 0


def _extract_price(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(" ", "").replace(",", ".")
    if not text:
        return 0.0
    m = re.search(r"-?\d+(?:\.\d+)?", text)
    if not m:
        return 0.0
    try:
        return float(m.group(0))
    except Exception:
        return 0.0


def _get_with_candidates(row: Dict[str, Any], header_map: Dict[str, str], direct: List[str], contains_all: List[Tuple[str, ...]]) -> Any:
    for key in direct:
        k = _norm_header(key)
        src = header_map.get(k)
        if src is not None:
            return row.get(src)

    # contains all keywords in normalized header
    normalized_keys = list(header_map.keys())
    for combo in contains_all:
        for hk in normalized_keys:
            if all(part in hk for part in combo):
                return row.get(header_map[hk])
    return None


def _build_required_columns(header_map: Dict[str, str]) -> Dict[str, str]:
    spec: Dict[str, Tuple[List[str], List[Tuple[str, ...]]]] = {
        "articul": (
            ["Артикул", "articul", "sku", "код товару"],
            [("артикул",), ("sku",), ("код", "товар")],
        ),
        "name": (
            ["Найменування", "Назва", "name"],
            [("наймен",), ("назва",), ("name",)],
        ),
        "brand": (
            ["Бренд", "brand"],
            [("бренд",), ("brand",)],
        ),
        "barcode": (
            ["Штрих код", "barcode", "штрихкод"],
            [("штрих",), ("barcode",)],
        ),
        "qty": (
            ["custom_NavnstLKtekst", "Наявність", "Доступно", "Склад"],
            [("custom", "navnstlktekst"), ("наявн",), ("доступ",), ("склад",)],
        ),
        "price": (
            ["userdiscountprice", "Ціна з персональною знижкою", "Ціна"],
            [("userdiscountprice",), ("персональ", "ціна"), ("discount", "price")],
        ),
        "termin": (
            ["custom_terminpridatnosti13", "Термін придатності"],
            [("custom", "terminpridatnosti13"), ("термін", "придат")],
        ),
    }

    found: Dict[str, str] = {}
    for field, (direct, contains_all) in spec.items():
        value = _get_with_candidates({k: k for k in header_map.values()}, header_map, direct, contains_all)
        if value is not None:
            found[field] = str(value)

    missing = [k for k in ["articul", "name", "qty", "price"] if k not in found]
    if missing:
        raise RuntimeError(
            f"Required export columns missing: {', '.join(missing)}. Available columns: {list(header_map.values())}"
        )
    return found


def _read_xlsx(path: Path) -> Tuple[List[str], List[Dict[str, Any]]]:
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise RuntimeError("Export xlsx is empty")
    header = [str(x or "").strip() for x in rows[0]]
    out: List[Dict[str, Any]] = []
    for r in rows[1:]:
        if not any(x not in (None, "") for x in r):
            continue
        row = {header[i]: r[i] if i < len(r) else None for i in range(len(header))}
        out.append(row)
    return header, out


def _read_xls(path: Path) -> Tuple[List[str], List[Dict[str, Any]]]:
    try:
        import pandas as pd
    except Exception as e:
        raise RuntimeError("pandas is required to parse .xls exports") from e

    try:
        df = pd.read_excel(path, engine="xlrd")
    except Exception as e:
        raise RuntimeError(f"Failed to parse .xls file: {e}") from e

    header = [str(c or "").strip() for c in list(df.columns)]
    out: List[Dict[str, Any]] = []
    for _, row in df.iterrows():
        values: Dict[str, Any] = {}
        empty = True
        for h in header:
            v = row.get(h)
            if v == v and v is not None and str(v).strip() != "":
                empty = False
            values[h] = None if (v != v) else v
        if empty:
            continue
        out.append(values)
    return header, out


def parse_export_to_json_records(export_path: Path) -> List[Dict[str, Any]]:
    suffix = export_path.suffix.lower()
    if suffix == ".xlsx":
        header, rows = _read_xlsx(export_path)
    elif suffix == ".xls":
        header, rows = _read_xls(export_path)
    else:
        raise RuntimeError(f"Unsupported export extension: {suffix}")

    header_map = {_norm_header(h): h for h in header if str(h).strip()}
    resolved = _build_required_columns(header_map)

    result: List[Dict[str, Any]] = []
    for row in rows:
        articul = str(row.get(resolved["articul"]) or "").strip()
        name = str(row.get(resolved["name"]) or "").strip()
        if not articul and not name:
            continue
        brand = str(row.get(resolved.get("brand", "")) or "").strip() if "brand" in resolved else ""
        barcode = str(row.get(resolved.get("barcode", "")) or "").strip() if "barcode" in resolved else ""
        qty = _extract_int_qty(row.get(resolved["qty"]))
        price = _extract_price(row.get(resolved["price"]))
        termin = str(row.get(resolved.get("termin", "")) or "").strip() if "termin" in resolved else ""

        result.append(
            {
                "articul": articul,
                "name": name,
                "brand": brand,
                "barcode": barcode,
                "qty": qty,
                "price": price,
                "termin": termin,
            }
        )
    return result


def _is_login_page(page) -> bool:
    url = (page.url or "").lower()
    if "/login" in url:
        return True
    try:
        if page.locator("input[type='password']").first.is_visible(timeout=1200):
            return True
    except Exception:
        pass
    return False


def _find_first_visible(page, selectors: List[str]):
    for s in selectors:
        loc = page.locator(s).first
        try:
            if loc.is_visible(timeout=1200):
                return loc
        except Exception:
            continue
    return None


def ensure_login(page, context) -> None:
    if not _is_login_page(page):
        return
    if not SUP2_USERNAME or not SUP2_PASSWORD:
        raise RuntimeError("Session expired and SUP2_USERNAME/SUP2_PASSWORD are not set")

    page.goto(SUP2_LOGIN_URL, wait_until="domcontentloaded", timeout=SUP2_TIMEOUT_MS)

    user_loc = _find_first_visible(
        page,
        [
            "input[name='username']",
            "input[name='login']",
            "input[name='email']",
            "input[type='email']",
            "input[id*='user']",
        ],
    )
    pass_loc = _find_first_visible(
        page,
        [
            "input[name='password']",
            "input[type='password']",
            "input[id*='pass']",
        ],
    )
    if user_loc is None or pass_loc is None:
        raise RuntimeError("Login form fields not found")

    user_loc.fill(SUP2_USERNAME, timeout=SUP2_TIMEOUT_MS)
    pass_loc.fill(SUP2_PASSWORD, timeout=SUP2_TIMEOUT_MS)

    submit = _find_first_visible(
        page,
        [
            "button[type='submit']",
            "input[type='submit']",
            "button:has-text('Увійти')",
            "button:has-text('Вхід')",
        ],
    )
    if submit is not None:
        submit.click(timeout=SUP2_TIMEOUT_MS)
    else:
        pass_loc.press("Enter")

    deadline = time.time() + 20
    while time.time() < deadline:
        if "/client/" in (page.url or "") and not _is_login_page(page):
            break
        page.wait_for_timeout(500)
    if _is_login_page(page):
        raise RuntimeError("Login verification failed")

    storage = Path(SUP2_STORAGE_STATE_FILE)
    if not storage.is_absolute():
        storage = ROOT / storage
    storage.parent.mkdir(parents=True, exist_ok=True)
    context.storage_state(path=str(storage))


def select_in_stock_filter(page) -> None:
    # First try a direct id pattern from real Dobavki markup:
    #   #select2-filtercustomNayavnsttovaru-*-container
    rendered = page.locator(
        "span.select2-selection__rendered[id^='select2-filtercustomNayavnsttovaru-'][id$='-container']"
    ).first
    if rendered.count() == 0:
        # Fallback by label text near select2
        rendered = page.locator(
            "xpath=//label[contains(., 'Наявність товару')]/following::span[contains(@class,'select2-selection__rendered')][1]"
        ).first
    if rendered.count() == 0:
        # Last fallback by any matching select id/name
        sel = page.locator("select[name*='Nayav'], select[name*='nayav'], select[id*='Nayav'], select[id*='nayav']").first
        if sel.count() > 0:
            sel_id = sel.get_attribute("id") or ""
            if sel_id:
                rendered2 = page.locator(f"#select2-{sel_id}-container").first
                if rendered2.count() > 0:
                    rendered = rendered2

    rendered.wait_for(state="visible", timeout=SUP2_TIMEOUT_MS)
    current = (rendered.inner_text(timeout=2000) or "").strip().lower()
    if "доступ" in current and "склад" in current:
        return

    selection = rendered.locator("xpath=ancestor::span[contains(@class,'select2-selection')][1]").first
    opened = False
    for click_mode in ("normal", "force", "js"):
        try:
            if click_mode == "normal":
                selection.click(timeout=3000)
            elif click_mode == "force":
                selection.click(timeout=3000, force=True)
            else:
                selection.evaluate("el => el.click()")
            page.locator(".select2-container--open").first.wait_for(state="visible", timeout=2500)
            opened = True
            break
        except Exception:
            continue
    if not opened:
        raise RuntimeError("Could not open select2 for 'Наявність товару'")

    search_input = page.locator(".select2-container--open input.select2-search__field").first
    search_input.wait_for(state="visible", timeout=SUP2_TIMEOUT_MS)
    search_input.fill("Доступно на складі", timeout=SUP2_TIMEOUT_MS)

    option_exact = page.locator(".select2-results__option", has_text="Доступно на складі").first
    if option_exact.count() > 0:
        try:
            option_exact.wait_for(state="visible", timeout=4000)
            option_exact.click(timeout=SUP2_TIMEOUT_MS)
        except Exception:
            search_input.press("Enter")
    else:
        search_input.press("Enter")

    page.wait_for_timeout(900)
    now = (rendered.inner_text(timeout=4000) or "").strip().lower()
    if not ("доступ" in now and "склад" in now):
        raise RuntimeError(f"Failed to set filter to 'Доступно на складі' (current={now!r})")


def trigger_export_download(page, export_dir: Path) -> Path:
    btn = page.get_by_role("button", name=re.compile(r"Експортувати", re.IGNORECASE)).first
    if btn.count() == 0:
        btn = page.locator("input[type='submit'][value*='Експортувати'], button:has-text('Експортувати')").first
    btn.wait_for(state="visible", timeout=SUP2_TIMEOUT_MS)

    with page.expect_download(timeout=SUP2_EXPORT_WAIT_SEC * 1000) as download_info:
        btn.click(timeout=SUP2_TIMEOUT_MS)
    download = download_info.value

    suggested = download.suggested_filename or "dobavki_export.xls"
    ext = Path(suggested).suffix.lower() or ".xls"
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    target = export_dir / f"dobavki_export_{ts}{ext}"
    download.save_as(str(target))

    if not target.exists() or target.stat().st_size <= 0:
        raise RuntimeError("Export download failed: file missing or empty")
    return target


def run_export() -> Dict[str, Any]:
    export_dir = Path(SUP2_EXPORT_DIR)
    if not export_dir.is_absolute():
        export_dir = ROOT / export_dir
    export_dir.mkdir(parents=True, exist_ok=True)

    storage_state = Path(SUP2_STORAGE_STATE_FILE)
    if not storage_state.is_absolute():
        storage_state = ROOT / storage_state

    credentials = Path(GDRIVE_CREDENTIALS_FILE)
    if not credentials.is_absolute():
        credentials = ROOT / credentials

    launch_kwargs: Dict[str, Any] = {"headless": SUP2_HEADLESS, "downloads_path": str(export_dir)}

    with sync_playwright() as pw:
        browser = pw.chromium.launch(**launch_kwargs)
        context_kwargs: Dict[str, Any] = {"accept_downloads": True}
        if storage_state.exists():
            context_kwargs["storage_state"] = str(storage_state)
        context = browser.new_context(**context_kwargs)
        page = context.new_page()

        page.goto(SUP2_EXPORT_URL, wait_until="domcontentloaded", timeout=SUP2_TIMEOUT_MS)
        ensure_login(page, context)

        if "/product/list" not in (page.url or ""):
            page.goto(SUP2_EXPORT_URL, wait_until="domcontentloaded", timeout=SUP2_TIMEOUT_MS)

        select_in_stock_filter(page)
        export_file = trigger_export_download(page, export_dir)

        browser.close()

    records = parse_export_to_json_records(export_file)
    json_path = export_dir / SUP2_EXPORT_JSON_NAME
    json_path.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")

    if not json_path.exists() or json_path.stat().st_size <= 0:
        raise RuntimeError("JSON save failed")

    if not GDRIVE_FOLDER_ID:
        raise RuntimeError("GDRIVE_FOLDER_ID is not set")
    if not credentials.exists():
        raise RuntimeError(f"GDRIVE_CREDENTIALS_FILE not found: {credentials}")

    drive_file_id = upload_or_update_json(
        credentials_file=credentials,
        folder_id=GDRIVE_FOLDER_ID,
        local_file=json_path,
        remote_filename=SUP2_EXPORT_JSON_NAME,
    )

    return {
        "ok": True,
        "rows": len(records),
        "json_path": str(json_path),
        "drive_uploaded": True,
        "drive_file_id": drive_file_id,
        "export_file": str(export_file),
        "size": json_path.stat().st_size,
    }


def main() -> int:
    try:
        summary = run_export()
        print(json.dumps(summary, ensure_ascii=False), flush=True)
        return 0
    except PWTimeoutError as e:
        out = {"ok": False, "stage": "playwright_timeout", "error": str(e)}
        print(json.dumps(out, ensure_ascii=False), flush=True)
        return 2
    except Exception as e:
        out = {"ok": False, "stage": "export", "error": f"{type(e).__name__}: {e}"}
        print(json.dumps(out, ensure_ascii=False), flush=True)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
