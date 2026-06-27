from __future__ import annotations

import argparse
import json
import os
import re
import sys
import urllib.error
import urllib.request
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Tuple

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

load_dotenv(ROOT / ".env")


ARTICLE_HEADER_CANDIDATES = {"Артикул"}
QTY_HEADER_CANDIDATES = {"Кіл-ст", "Кіль-ст", "Кількість", "Кіл-сть"}
PRICE_HEADER_CANDIDATES = {"Ціна з макс. Знижкою", "Цена с макс. скидкой", "Ціна з макс. знижкою"}


@dataclass(frozen=True)
class SportAtletItem:
    article: str
    qty: int
    name: str = ""


def _env(name: str, default: str = "") -> str:
    return (os.getenv(name) or default).strip()


def _to_bool(value: str, default: bool = False) -> bool:
    raw = (value or "").strip().lower()
    if not raw:
        return default
    return raw in {"1", "true", "yes", "y", "on"}


def _parse_article(description: str) -> str:
    text = str(description or "").strip()
    if not text:
        return ""
    return re.split(r"[\s,]+", text, maxsplit=1)[0].strip()


def parse_sportatlet_items(order: Dict[str, Any]) -> List[Dict[str, Any]]:
    products = order.get("products") or []
    merged: Dict[str, Dict[str, Any]] = {}
    for p in products:
        if not isinstance(p, dict):
            continue
        article = _parse_article(str(p.get("description") or ""))
        if not article:
            continue
        try:
            qty = int(p.get("amount") or 1)
        except Exception:
            qty = 1
        if qty < 1:
            qty = 1
        name = str(p.get("text") or p.get("name") or "").strip()
        if article in merged:
            merged[article]["qty"] = int(merged[article].get("qty") or 0) + qty
            if not merged[article].get("name") and name:
                merged[article]["name"] = name
        else:
            merged[article] = {"article": article, "qty": qty, "name": name}
    out = list(merged.values())
    if not out:
        raise RuntimeError("No products for Sport-atlet order")
    return out


def build_sportatlet_salesdrive_products(order: Dict[str, Any]) -> List[Dict[str, Any]]:
    products = order.get("products") or []
    out: List[Dict[str, Any]] = []
    for p in products:
        if not isinstance(p, dict):
            continue
        raw_id = p.get("id")
        if raw_id in (None, ""):
            continue
        desc = str(p.get("description") or "").strip()
        out.append(
            {
                "id": raw_id,
                "name": p.get("text") or p.get("name") or "",
                "costPerItem": p.get("costPerItem"),
                "amount": p.get("amount"),
                "description": desc,
                "discount": p.get("discount") if p.get("discount") is not None else "",
                "sku": _parse_article(desc),
            }
        )
    return out


def sportatlet_template_url() -> str:
    return _env("SPORTATLET_TEMPLATE_URL", "https://sport-atlet.ua/opt/Prays-opt-sport-atlet.xlsx")


def sportatlet_tg_bot_token() -> str:
    token = _env("SPORTATLET_TG_BOT_TOKEN") or _env("TG_BOT_TOKEN")
    if not token:
        raise RuntimeError("SPORTATLET_TG_BOT_TOKEN (or TG_BOT_TOKEN) is empty")
    return token


def sportatlet_tg_chat_id() -> str:
    chat_id = _env("SPORTATLET_TG_CHAT_ID")
    if not chat_id:
        raise RuntimeError("SPORTATLET_TG_CHAT_ID is empty")
    return chat_id


def sportatlet_dry_run_enabled() -> bool:
    return _to_bool(_env("SPORTATLET_DRY_RUN", "0"), False)


def sportatlet_number_sup_value() -> str:
    return _env("SPORTATLET_NUMBERSUP_VALUE", "TG_SENT")


def _files_dir() -> Path:
    raw = _env("SPORTATLET_FILES_DIR", "sportatlet_files")
    p = Path(raw)
    if not p.is_absolute():
        p = ROOT / p
    p.mkdir(parents=True, exist_ok=True)
    return p


def find_header_row_and_cols(ws, max_scan_rows: int = 50, max_scan_cols: int = 60) -> Tuple[int | None, int | None, int | None, int | None]:
    max_r = min(ws.max_row, max_scan_rows)
    max_c = min(ws.max_column, max_scan_cols)

    for r in range(1, max_r + 1):
        art_col = None
        qty_col = None
        price_col = None
        for c in range(1, max_c + 1):
            v = ws.cell(r, c).value
            if not isinstance(v, str):
                continue
            s = v.strip()
            if s in ARTICLE_HEADER_CANDIDATES:
                art_col = c
            if s in QTY_HEADER_CANDIDATES:
                qty_col = c
            if s in PRICE_HEADER_CANDIDATES:
                price_col = c
        if art_col and qty_col:
            return r, art_col, qty_col, price_col
    return None, None, None, None


def build_article_to_row_index(ws, header_row: int, art_col: int) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        v = ws.cell(r, art_col).value
        if isinstance(v, str):
            article = v.strip()
            if article:
                mapping[article] = r
    return mapping


def download_template_xlsx(url: str | None = None) -> bytes:
    target_url = (url or sportatlet_template_url()).strip()
    if not target_url:
        raise RuntimeError("SPORTATLET_TEMPLATE_URL is empty")
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/121.0.0.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "uk-UA,uk;q=0.9,en-US;q=0.8,en;q=0.7,ru;q=0.6",
        "Referer": "https://sport-atlet.com.ua/",
        "Connection": "keep-alive",
    }
    resp = requests.get(target_url, headers=headers, timeout=60, allow_redirects=True)
    resp.raise_for_status()
    if not resp.content:
        raise RuntimeError("Sport-atlet template download returned empty response")
    return resp.content


def fill_sportatlet_quantities(xlsx_bytes: bytes, items: List[Dict[str, Any]]) -> Tuple[bytes, Dict[str, str]]:
    wb_write = load_workbook(BytesIO(xlsx_bytes))
    ws_write = wb_write.active
    wb_values = load_workbook(BytesIO(xlsx_bytes), data_only=True)
    ws_values = wb_values.active

    header_row, art_col, qty_col, price_col = find_header_row_and_cols(ws_write)
    if not header_row or not art_col or not qty_col:
        raise RuntimeError("Не найден заголовок с колонками 'Артикул' и 'Кіл-ст' в файле Sport-atlet.")

    index = build_article_to_row_index(ws_write, header_row, art_col)
    missing: List[str] = []
    price_map: Dict[str, str] = {}

    for it in items:
        article = str(it.get("article") or "").strip()
        if not article:
            continue
        row = index.get(article)
        if not row:
            missing.append(article)
            continue
        try:
            qty = int(it.get("qty") or 1)
        except Exception:
            qty = 1
        if qty < 1:
            qty = 1
        ws_write.cell(row, qty_col).value = qty
        if price_col:
            pv = ws_values.cell(row, price_col).value
            if pv is not None:
                try:
                    price_map[article] = f"{round(float(pv), 2):.2f}"
                except Exception:
                    price_map[article] = str(pv)

    if missing:
        raise RuntimeError("Артикулы не найдены в шаблоне Sport-atlet: " + ", ".join(missing))

    out = BytesIO()
    wb_write.save(out)
    out.seek(0)
    return out.read(), price_map


def build_sportatlet_xlsx_file(ttn: str, items: List[Dict[str, Any]], xlsx_bytes: bytes | None = None) -> Path:
    ttn_clean = str(ttn or "").strip()
    if not ttn_clean:
        raise RuntimeError("TTN is empty")
    src = xlsx_bytes if xlsx_bytes is not None else download_template_xlsx()
    out_bytes, _ = fill_sportatlet_quantities(src, items)
    out_path = _files_dir() / f"{ttn_clean}.xlsx"
    out_path.write_bytes(out_bytes)
    if out_path.stat().st_size <= 0:
        raise RuntimeError("Generated Sport-atlet XLSX is empty")
    return out_path


def download_sportatlet_label_pdf(ttn: str) -> Path:
    api_key = _env("BIOTUS_NP_API_KEY") or _env("NP_API_KEY")
    if not api_key:
        raise RuntimeError("BIOTUS_NP_API_KEY (or NP_API_KEY) is not set")

    ttn_clean = str(ttn or "").strip()
    if not ttn_clean:
        raise RuntimeError("TTN is empty")
    out_path = _files_dir() / f"marking-{ttn_clean}.pdf"
    url = (
        "https://my.novaposhta.ua/orders/printMarking100x100/"
        f"orders[]/{ttn_clean}/type/pdf/apiKey/{api_key}/zebra"
    )
    try:
        req = urllib.request.Request(url, method="GET")
        with urllib.request.urlopen(req, timeout=25) as resp:
            status = int(getattr(resp, "status", 200) or 200)
            if status >= 400:
                raise RuntimeError(f"Nova Poshta API returned status {status}")
            data = resp.read()
            if not data:
                raise RuntimeError("Downloaded PDF is empty")
            out_path.write_bytes(data)
    except urllib.error.HTTPError as e:
        raise RuntimeError(f"NP API HTTP {e.code}: {e.reason}") from e
    except urllib.error.URLError as e:
        raise RuntimeError(f"NP API connection error: {e}") from e

    if out_path.stat().st_size <= 0:
        raise RuntimeError("Label download failed: file size is zero")
    return out_path


def send_sportatlet_telegram_media_group(token: str, chat_id: str, xlsx_path: Path, pdf_path: Path, caption: str) -> None:
    url = f"https://api.telegram.org/bot{token}/sendMediaGroup"
    media = [
        {
            "type": "document",
            "media": "attach://xlsx_file",
            "caption": str(caption or "").strip(),
        },
        {
            "type": "document",
            "media": "attach://pdf_file",
        },
    ]
    with xlsx_path.open("rb") as xlsx_fh, pdf_path.open("rb") as pdf_fh:
        files = {
            "xlsx_file": (xlsx_path.name, xlsx_fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            "pdf_file": (pdf_path.name, pdf_fh, "application/pdf"),
        }
        resp = requests.post(
            url,
            data={"chat_id": chat_id, "media": json.dumps(media, ensure_ascii=False)},
            files=files,
            timeout=60,
        )
    if resp.status_code < 200 or resp.status_code >= 300:
        raise RuntimeError(f"Telegram sendMediaGroup HTTP {resp.status_code}: {resp.text[:300]}")
    try:
        payload = resp.json()
    except Exception as e:
        raise RuntimeError(f"Telegram sendMediaGroup returned non-JSON response: {resp.text[:300]}") from e
    if not bool(payload.get("ok")):
        raise RuntimeError(f"Telegram sendMediaGroup failed: {payload}")


def _extract_ttn_from_order(order: Dict[str, Any]) -> str:
    odd = order.get("ord_delivery_data") or []
    d0 = odd[0] if isinstance(odd, list) and odd else (odd if isinstance(odd, dict) else {})
    return str((d0 or {}).get("trackingNumber") or "").strip()


def run_sportatlet_telegram_flow(order: Dict[str, Any], ttn: str) -> Dict[str, Any]:
    items = parse_sportatlet_items(order)
    token = sportatlet_tg_bot_token()
    chat_id = sportatlet_tg_chat_id()

    if sportatlet_dry_run_enabled():
        return {
            "ok": True,
            "supplier": "sportatlet",
            "ttn": ttn,
            "chat_id": chat_id,
            "items": items,
            "xlsx_path": "",
            "pdf_path": "",
            "numberSup": sportatlet_number_sup_value(),
            "dry_run": True,
        }

    xlsx_path = build_sportatlet_xlsx_file(ttn, items)
    pdf_path = download_sportatlet_label_pdf(ttn)
    send_sportatlet_telegram_media_group(
        token=token,
        chat_id=chat_id,
        xlsx_path=xlsx_path,
        pdf_path=pdf_path,
        caption=str(ttn).strip(),
    )
    return {
        "ok": True,
        "supplier": "sportatlet",
        "ttn": ttn,
        "chat_id": chat_id,
        "items": items,
        "xlsx_path": str(xlsx_path),
        "pdf_path": str(pdf_path),
        "numberSup": sportatlet_number_sup_value(),
    }


def _load_order_payload(order_json_arg: str, order_json_file: str) -> Dict[str, Any]:
    if order_json_file:
        p = Path(order_json_file)
        obj = json.loads(p.read_text(encoding="utf-8"))
        if not isinstance(obj, dict):
            raise RuntimeError("order-json-file must contain JSON object")
        return obj
    raw_arg = (order_json_arg or "").strip()
    if not raw_arg:
        raise RuntimeError("--order-json or --order-json-file is required")
    p2 = Path(raw_arg)
    if p2.exists() and p2.is_file():
        obj = json.loads(p2.read_text(encoding="utf-8"))
    else:
        obj = json.loads(raw_arg)
    if not isinstance(obj, dict):
        raise RuntimeError("order-json must be a JSON object")
    return obj


def main() -> int:
    ap = argparse.ArgumentParser(description="Sport-atlet supplier 43 Telegram flow")
    ap.add_argument("--order-json", default="", help="Order JSON string or path to JSON file")
    ap.add_argument("--order-json-file", default="", help="Path to order JSON file")
    ap.add_argument("--ttn", default="", help="TTN override")
    args = ap.parse_args()

    try:
        order = _load_order_payload(args.order_json, args.order_json_file)
        ttn = (args.ttn or "").strip() or _extract_ttn_from_order(order)
        if not ttn:
            raise RuntimeError("TTN is empty (pass --ttn or provide ord_delivery_data[0].trackingNumber)")
        result = run_sportatlet_telegram_flow(order, ttn)
        print(json.dumps(result, ensure_ascii=False))
        return 0 if bool(result.get("ok")) else 1
    except Exception as e:
        order = {}
        try:
            order = _load_order_payload(args.order_json, args.order_json_file)
        except Exception:
            pass
        items: List[Dict[str, Any]] = []
        try:
            if order:
                items = parse_sportatlet_items(order)
        except Exception:
            items = []
        ttn = (args.ttn or "").strip() or _extract_ttn_from_order(order)
        fail = {
            "ok": False,
            "supplier": "sportatlet",
            "reason": str(e),
            "ttn": ttn,
            "items": items,
        }
        print(json.dumps(fail, ensure_ascii=False))
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
